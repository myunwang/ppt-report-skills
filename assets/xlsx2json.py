"""
xlsx2json.py — 把 Excel/CSV 转成 report-deck 用的 JSON。

可作为独立调试工具单跑，也被 build.py 自动调用。

约定
----
1. 每个 slide 对应一个文件：
   - src/data/slide-N.xlsx   多 sheet：每个 sheet → JSON 顶层一个 key
   - src/data/slide-N.csv    单表：整份 → JSON 顶层一个 array
   - src/data/slide-N.json   原生 JSON：直接读取

2. 每个 sheet / csv 的格式：
   第 1 行 = 表头，后续行 = 数据
   转换为 [ {col1: val, col2: val}, ... ]
   单元格中的数字自动转 number；其他保留字符串

3. 以 `_` 开头的 sheet 名或列名会被跳过（用作备注、辅助计算等）

用法
----
    # 单文件转换（输出到 stdout）
    python3 xlsx2json.py src/data/slide-3.xlsx

    # 转换并写入
    python3 xlsx2json.py src/data/slide-3.xlsx > src/data/slide-3.json

    # 转换整个 data/ 目录
    python3 xlsx2json.py src/data/

依赖
----
- xlsx 支持：openpyxl（pip install openpyxl）
- csv / json：标准库

若 openpyxl 未安装，仅 csv / json 可用，xlsx 文件会被跳过并提示。
"""
from __future__ import annotations
import csv
import json
import sys
from pathlib import Path


def _coerce(v):
    """单元格值类型转换：数字 → number；空 → None；其他 → str."""
    if v is None:
        return None
    if isinstance(v, (int, float, bool)):
        return v
    s = str(v).strip()
    if s == '':
        return None
    # 尝试 int
    try:
        if '.' not in s and 'e' not in s.lower():
            return int(s)
    except ValueError:
        pass
    # 尝试 float
    try:
        return float(s)
    except ValueError:
        pass
    return s


def _rows_to_objects(rows):
    """[[header...], [row...], ...] → [{col: val}, ...]，过滤掉 _ 开头的列."""
    if not rows:
        return []
    header = [str(h).strip() if h is not None else '' for h in rows[0]]
    keep_cols = [i for i, h in enumerate(header) if h and not h.startswith('_')]
    out = []
    for row in rows[1:]:
        # 整行空则跳过
        if all(v is None or str(v).strip() == '' for v in row):
            continue
        obj = {}
        for i in keep_cols:
            if i < len(row):
                obj[header[i]] = _coerce(row[i])
            else:
                obj[header[i]] = None
        out.append(obj)
    return out


def xlsx_to_dict(path: Path) -> dict:
    """xlsx → {sheet_name: [rows]}，跳过 _ 开头的 sheet."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError(
            f'读取 {path.name} 需要 openpyxl。安装：pip install openpyxl'
        )
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith('_'):
            continue
        ws = wb[sheet_name]
        rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
        result[sheet_name] = _rows_to_objects(rows)
    wb.close()
    return result


def csv_to_list(path: Path) -> list:
    """csv → [{col: val}, ...]."""
    with path.open('r', encoding='utf-8-sig', newline='') as f:
        reader = csv.reader(f)
        rows = list(reader)
    return _rows_to_objects(rows)


def convert_file(path: Path):
    """根据扩展名分派转换；返回 Python 对象或 None（跳过）."""
    suffix = path.suffix.lower()
    if suffix == '.json':
        return json.loads(path.read_text(encoding='utf-8'))
    if suffix == '.xlsx':
        return xlsx_to_dict(path)
    if suffix == '.csv':
        return csv_to_list(path)
    return None


def convert_dir(d: Path) -> dict:
    """遍历目录，返回 {slide_n: object} 映射。优先级 json > xlsx > csv（同 slide 编号）."""
    by_slide = {}  # n -> (priority, path)
    PRIORITY = {'.json': 0, '.xlsx': 1, '.csv': 2}
    for p in sorted(d.iterdir()):
        if not p.is_file() or p.name.startswith('.') or p.name.startswith('_'):
            continue
        if p.suffix.lower() not in PRIORITY:
            continue
        # 解析 slide-N.xxx
        stem = p.stem
        if not stem.startswith('slide-'):
            continue
        try:
            n = int(stem[len('slide-'):])
        except ValueError:
            continue
        prio = PRIORITY[p.suffix.lower()]
        if n not in by_slide or prio < by_slide[n][0]:
            by_slide[n] = (prio, p)

    out = {}
    for n, (_, p) in sorted(by_slide.items()):
        try:
            obj = convert_file(p)
            if obj is not None:
                out[n] = (p, obj)
        except Exception as e:
            print(f'⚠ {p.name}: {e}', file=sys.stderr)
    return out


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    target = Path(sys.argv[1])
    if not target.exists():
        print(f'不存在：{target}', file=sys.stderr)
        sys.exit(1)

    if target.is_file():
        obj = convert_file(target)
        if obj is None:
            print(f'不支持的文件类型：{target}', file=sys.stderr)
            sys.exit(1)
        json.dump(obj, sys.stdout, ensure_ascii=False, indent=2)
        sys.stdout.write('\n')
    else:
        result = convert_dir(target)
        for n, (path, obj) in result.items():
            print(f'# slide-{n} ({path.name})')
            json.dump(obj, sys.stdout, ensure_ascii=False, indent=2)
            sys.stdout.write('\n\n')


if __name__ == '__main__':
    main()
